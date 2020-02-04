from openpyxl import load_workbook
import numpy as np
from main import Country


class Reader:
    countries = np.zeros((10, 10))
    wb = load_workbook('CountryData.xlsx')  # opens the Excel file
    source = wb.active
    ws = wb.copy_worksheet(source)


def country_init(n):
    ws = Reader.wb.copy_worksheet(Reader.source)
    c = Country.Country()
    s = input("Enter country " + str(n) + "\n")
    for i in range(ws.min_row + 1, ws.max_row):
        country = ws.cell(row=i, column=1)
        if country.value == s:
            c.ID_TAG = ws.cell(row=i, column=2).value
            c.GDP = ws.cell(row=i, column=3).value
            c.ARMY_ASSETS = ws.cell(row=i, column=4).value
            c.NAVY_ASSETS = ws.cell(row=i, column=5).value
            c.AIR_FORCE_ASSETS = ws.cell(row=i, column=6).value
            c.MANPOWER = ws.cell(row=i, column=7).value
            c.NUCLEAR_CAPABILITIES = ws.cell(row=i, column=8).value
    return c
