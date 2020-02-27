from openpyxl import load_workbook
import numpy as np

wb = load_workbook('Coefficients.xlsx')
source = wb.active
ws = wb.copy_worksheet(source)

a = np.zeros(ws.max_row, ws.max_column)
b = np.zeros(ws.max_row, 1)

for i in range(1,ws.max_row):
    for j in range(1,ws.max_column-2):
        a[i][j] = ws.cell(row=i, column=j).value

for i in range(1,ws.max_row):
    b[i][1] = ws.cell(row=i, column=1).value

x = np.linalg.solve(a, b)

print(x)

