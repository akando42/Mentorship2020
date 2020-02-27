import numpy as np
from openpyxl import load_workbook
from main import Country
from main import Reader
from main import Reader2
from main import Coordinates


class Main(object):
    def __init__(self):
        n = 4
    def prob(self):
        p = (self.ARMYc*self.armyCommitments) + (self.NAVYc*self.navyCommitments) + (self.AFc*self.afCommitments) + (self.NC*self.ncCommitments)
        p *= self.maneuvers[self.move_player][self.move_comp]
        return p
    def winner(self,a,b):
            return 100*(a/b)-20
    def getUserMoves(self,n):
        n = 1
        military_type = input("Entering the following number for:\n[0] Offense\n[1] Defense\n")
        if (military_type == '0'):
            maneuver = input("Pick one of the following offensive maneuvers: \n[0] Push Middle\n[1] Flank Right\n[2] Flank Left\n[3] Pincer\n[4] Push Equally\n[5]Trap on Two Fronts\n[999] Go back to beginning\n")
            if (maneuver == '0'):
                print('You tried to push middle\n')
                self.move_player = 1
                self.player_str = "pushed middle"
            elif (maneuver == '1'):
                print('You tried to flanked right\n')
                self.move_player = 2
                self.player_str = "flank right"
            elif (maneuver == '2'):
                print('You tried to flanked left\n')
                self.move_player = 3
                self.player_str = "flanked left"
            elif (maneuver == '3'):
                print('You tried to do a pincer movement\n')
                self.move_player = 4
                self.player_str = "used a pincer movement"
            elif (maneuver == '4'):
                print('You tried to push equally on all fronts\n')
                self.move_player = 5
                self.player_str = "pushed equally on all fronts"
            elif (maneuver == '5'):
                print('You tried to trap the enemy on two fronts\n')
                self.move_player = 6
                self.player_str = "tried to trap the enemy on two fronts"
            elif (maneuver == '999'):
                print("Return to beginning")
            else:
                print('Wrong input. Try again.')
        if (military_type == '1'):
            maneuver = input(
                'Pick one of the following defensive maneuvers:\n[0] Retreat\n[1] Retreating Fire\n[2] Find and Hold a Choke Point\n[3] Turtle')
            if (maneuver == '0'):
                print('You tried to retreat\n')
                self.move_player = 7
                self.player_str = "tried to retreat"
            elif (maneuver == '1'):
                print('You tried to retreat to cover fire\n')
                self.move_player = 8
                self.player_str = "retreated with cover fire"
            elif (maneuver == '2'):
                print('You tried to find and hold a choke point\n')
                self.move_player = 9
                self.player_str = "held a choke point"
            elif (maneuver == '3'):
                print('You tried to create a defensive turtle for a better defense\n')
                self.move_player = 10
                self.player_str = "created a defensive turtle"
            elif (maneuver == '999'):
                print("Return to beginning")
            else:
                print('Wrong input. Try again.')

    def getArmyResources(self,c):
        ARMY = input("Enter number of army assets between 0 and " + str(c.ARMY_ASSETS) + "\n")
        if(int(ARMY) < c.ARMY_ASSETS):
            return int(ARMY)
    def getNavyResources(self,c):
        NAVY = input("Enter number of navy assets between 0 and " + str(c.NAVY_ASSETS) + "\n")
        if(int(NAVY) < c.NAVY_ASSETS):
            return int(NAVY)
    def getAFResources(self,c):
        AF = input("Enter number of air force assets between 0 and " + str(c.AIR_FORCE_ASSETS) + "\n")
        if(int(AF) < c.AIR_FORCE_ASSETS):
            return int(AF)
    def getNCResources(self,c):
        NC = input("Enter number of nuclear assets between 0 and " + str(c.NUCLEAR_CAPABILITIES) + "\n")
        if(int(NC) < c.NUCLEAR_CAPABILITIES):
            return int(NC)
    def determineComputerMove(self,n,C):
        n+= 100
        result = [0,0,0,0]
        #Put in some algorithm to determine smallest number of resources need to win a battle
        self.armyCommitments /= self.maneuvers[self.move_comp][self.move_player]
        result [0] = int(self.armyCommitments) + 15
        self.navyCommitments /= self.maneuvers[self.move_comp][self.move_player]
        result[1] = int(self.navyCommitments) + 15
        self.afCommitments /= self.maneuvers[self.move_comp][self.move_player]
        result[2] = int(self.afCommitments) + 15
        result[3] = 0

        return result
    # Game coefficients (subject to change as testing goes on)
    GDPc = 1
    ARMYc = .542
    NAVYc = .459
    AFc = .515
    MPc = 1
    NC = 0
    # Defines variable committed resources
    armyCommitments = 0
    navyCommitments = 0
    afCommitments = 0
    ncCommitments = 0
    # Sets up maneuver matrix
    maneuvers = np.zeros((11, 11))
    wb = load_workbook('MilitaryTactics.xlsx')
    source = wb.active
    ws = wb.copy_worksheet(source)
    for i in range(1, 11):
        for j in range(1, 11):
            maneuvers[i][j] = ws.cell(row=i, column=j).value
    # Records moves of both players to determine win probability
    move_player = 0
    move_comp = 0
    player_str = " "

class Main:

    M1 = Main()
    #Starts the game
    M1.move_comp = 1
    end_game = False
    comp_str = ""
    # Countries Init
    print('Welcome to the the Battle Simulator')
    c1 = Reader.country_init(1)
    c2 = Reader.country_init(2)
    print("You have chosen to be " + c1.ID_TAG)
    print("Your opponent is " + c2.ID_TAG)
    # Battle Location Init
    print("Select Battle Location")
    battle_loc = Coordinates.Cords()
    loc_type = input("Entering the following number for:\n[0] Custom Location\n[1] Random Location\n")
    if loc_type == 0:
        x = input("Enter x value:\n")
        y = input("Enter y value:\n")
        battle_loc.set_loc(x,y)
    else:
        battle_loc.set_rand_cords()
    print("Battle will take place at: " + str(battle_loc.get_loc()))
    while (end_game == False):
        #Gets user move
        move_player = M1.getUserMoves(1)
        if(M1.move_player == 1):
            M1.move_comp = 4
            comp_str = "used a pincer movement"
        if (M1.move_player == 2):
            M1.move_comp = 1
            comp_str = "pushed middle"
        if (M1.move_player == 3):
            M1.move_comp = 1
            comp_str = "pushed middle"
        if (M1.move_player == 4):
            M1.move_comp = 5
            comp_str = "pushed equally on all fronts"
        if (M1.move_player == 5):
            M1.move_comp = 8
            comp_str = "retreated with cover fire"
        if (M1.move_player == 6):
            M1.move_comp = 5
            comp_str = "pushed equally across all fronts"
        if (M1.move_player == 7):
            M1.move_comp = 6
            comp_str = "attacked on two fronts"
        if (M1.move_player ==8):
            M1.move_comp = 6
            comp_str = "used a pincer momvement"
        if (M1.move_player == 9):
            M1.move_comp = 6
            comp_str = "trapped you on two fronts"
        if (M1.move_player == 10):
            M1.move_comp = 2
            comp_str = "flanked your position"
        #Gets how much the user wants to commit
        M1.armyCommitments = M1.getArmyResources(c1)
        M1.navyCommitments = M1.getNavyResources(c1)
        M1.afCommitments = M1.getAFResources(c1)
        M1.ncCommitments = M1.getNCResources(c1)
        print(" You committed " + str(M1.armyCommitments) + " ground units\n",
              "You committed " + str(M1.navyCommitments) + " naval units\n",
              "You committed " + str(M1.afCommitments) + " air units\n",
              "You committed " + str(M1.ncCommitments) + " nuclear missiles\n",
              "")
        if(M1.ncCommitments != 0):
            print("No one won because the utilization of nuclear weapons causes mutually assured destruction")
            end_game = True
            break
        #number used to determine players chance of winning
        p1 = M1.prob()
        print("The computer " + comp_str + "\n")
        comp_resources = M1.determineComputerMove(p1,c2)
        print(" The computer committed " + str(comp_resources[0]) + " ground units\n",
              "The computer committed " + str(comp_resources[1]) + " naval units\n",
              "The computer committed " + str(comp_resources[2]) + " air units\n",
              "The computer committed " + str(comp_resources[3]) + " nuclear missiles\n",
              "")
        #player win percentage
        p2 = (M1.prob()/M1.maneuvers[M1.move_player][M1.move_comp]) * M1.maneuvers[M1.move_comp][M1.move_player]
        winner = M1.winner(p1,p2)
        print("There is an " + str(int(winner)) +" percent chance of you winning")
        print("The computer " + comp_str + " while you " + M1.player_str)
        end_game = True