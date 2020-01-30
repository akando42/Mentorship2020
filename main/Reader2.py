from openpyxl import load_workbook
import numpy as np

class Reader2:
    maneuvers = np.zeros((11, 11))
    wb = load_workbook('MilitaryTactics.xlsx')
    source = wb.active
    ws = wb.copy_worksheet(source)

    for i in range(1,11):
        for j in range(1,11):
            maneuvers[i][j] = ws.cell(row=i,column=j).value